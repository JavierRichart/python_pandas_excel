from pathlib import Path
from typing import Iterable, Optional
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
import shutil
import sys

import pandas as pd

APP_NAME = "excel-runner"
REQUIRED_HEADERS = ("Broker", "Broker_fee_total")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger(APP_NAME)


def base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def find_latest_xlsx(folder: Path) -> Path | None:
    files = [p for p in folder.glob(".xlsx") if not p.name.startswith("~$")]
    if not files:
        return None
    return max(files, key=lambda p: p.stat().st_mtime)


def validate_required_columns(df: pd.DataFrame,
                              required: Iterable[str] = ("Broker", "Broker_fee", "Quantity"), ) -> None:
    normalized = {str(c).strip().lower(): c for c in df.columns}
    missing = [col for col in required if col.lower() not in normalized]
    if missing:
        raise ValueError(
            f"Faltan columnas requeridas: {missing}. "
        )


def prepare_broker_table(
        df: pd.DataFrame,
        order=("Broker", "Broker_fee", "Quantity"),
        result_col="Broker_fee_total",
        keep_others: bool = True,
) -> pd.DataFrame:
    validate_required_columns(df, required=order)
    lower_to_orig = {str(c).strip().lower(): c for c in df.columns}
    renames = {lower_to_orig[col.lower()]: col for col in order}
    out = df.copy().rename(columns=renames)

    out["Broker"] = out["Broker"].astype(str).str.strip()

    out["Broker_fee"] = pd.to_numeric(
        out["Broker_fee"].astype(str).str.replace(",", ".", regex=False),
        errors="coerce"
    ).fillna(0)

    out["Quantity"] = pd.to_numeric(
        out["Quantity"].astype(str).str.replace(",", ".", regex=False),
        errors="coerce"
    )

    out[result_col] = out["Quantity"] * out["Broker_fee"]

    leading = list(order) + [result_col]
    if keep_others:
        tail = [c for c in out.columns if c not in leading]
        out = out[leading + tail]
    else:
        out = out[leading]

    return out


def normalize(s) -> str:
    return str(s).strip().lower() if s is not None else ""


def quote_sheet(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def add_broker_summary_sheet(xlsx_path: Path,
                             data_sheet: Optional[str] = None,
                             summary_sheet: str = "Broker_Summary") -> None:
    wb = load_workbook(xlsx_path)
    ws_data = None

    if data_sheet and data_sheet in wb.sheetnames:
        ws_data = wb[data_sheet]
    else:
        for ws in wb.workwheets:
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val is None:
                    continue
                headers[normalize(val)] = (col, str(val))
            if all(h.lower() in headers for h in (normalize(h) for h in REQUIRED_HEADERS)):
                ws_data = ws
                break
    if ws_data is None:
        raise ValueError(
            f"{REQUIRED_HEADERS} not found"
        )
    header_map = {}
    for col in range(1, ws_data.max_column + 1):
        val = ws_data.cell(row=1, column = col).value
        if val is None:
            continue
        header_map[normalize(val)] = col

    col_broker = header_map[normalize("Broker")]
    col_total = header_map[normalize("Broker_fee_total")]

    last_row = ws_data.max_row
    while last_row > 1 and ws_data.cell(row=last_row, column=col_broker).value in (None, ""):
        last_row = last_row - 1
    if last_row <= 1:
        raise ValueError("No rows")

    brokers = OrderedDict()
    for r in range(2, last_row + 1):
        b = ws_data.cell(row=r, column=col_broker).value
        if b is None or str(b).strip() == "":
            continue
        brokers.setdefault(str(b).strip(), True)

    if summary_sheet in wb.sheetnames:
        del wb[summary_sheet]
    ws_sum = wb.create_sheet(title=summary_sheet)

    ws_sum["A1"] = "Broker"
    ws_sum["B1"] = "Total_Broker_fee"

    sheet_ref = quote_sheet(ws_data.title)
    broker_col_letter = get_column_letter(col_broker)
    total_col_letter = get_column_letter(col_total)
    criteria_col = "A"

    data_broker_range = f"{sheet_ref}!{broker_col_letter}2:{broker_col_letter}{last_row}"
    data_total_range = f"{sheet_ref}!{total_col_letter}2:{total_col_letter}{last_row}"

    row = 2
    for broker in brokers.keys():
        ws_sum.cell(row=row, column=1, value=broker)
        formula = f"SUMIF({data_broker_range}, {criteria_col}{row}, {data_total_range}"
        ws_sum.cell(row=row, column=2, value=formula)
        row += 1

    wb.save(xlsx_path)